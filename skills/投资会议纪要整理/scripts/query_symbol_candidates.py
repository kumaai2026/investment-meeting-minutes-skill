#!/usr/bin/env python3
"""Query local stock-symbol resources and return ranked candidates."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from dataclasses import asdict, dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable

DEFAULT_SYMBOL_ROOT = Path("/Users/kumaai/Documents/Codex/workspace/投资纪要工作流/03 Resources/market-symbols")
DEFAULT_ALIAS_PATH = DEFAULT_SYMBOL_ROOT / "company_aliases.csv"


@dataclass
class Candidate:
    symbol: str
    name: str
    market: str
    confidence: float
    match_type: str
    source: str


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", "", value).strip().upper()


def normalize_code(value: str) -> str:
    return value.strip().upper().replace(" ", "")


def load_a_share(root: Path) -> list[dict[str, str]]:
    path = root / "a_share_list.csv"
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [
            {
                "symbol": row.get("symbol", "").strip(),
                "code": row.get("code", "").strip(),
                "name": row.get("name", "").strip(),
                "market": "A",
                "source": str(path),
            }
            for row in csv.DictReader(handle)
            if row.get("symbol") and row.get("name")
        ]


def load_us(root: Path) -> list[dict[str, str]]:
    source_dir = root / "global-stock-symbols-source"
    files = [
        ("NASDAQ", source_dir / "nasadaq_3105.csv"),
        ("NYSE", source_dir / "nyse_3105.csv"),
        ("AMEX", source_dir / "amex_3105.csv"),
    ]
    rows: list[dict[str, str]] = []
    for market, path in files:
        if not path.exists():
            continue
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                symbol = (row.get("Symbol") or "").strip()
                name = (row.get("Name") or "").strip()
                if symbol and name:
                    rows.append({"symbol": symbol, "code": symbol, "name": name, "market": market, "source": str(path)})
    return rows


def load_hk(root: Path) -> list[dict[str, str]]:
    path = root / "global-stock-symbols-source/HongLongListOfSecurities_300625.xlsx"
    if not path.exists():
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    rows: list[dict[str, str]] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    header_seen = False
    for values in sheet.iter_rows(values_only=True):
        if not header_seen:
            header_seen = values and values[0] == "Stock Code"
            continue
        code = str(values[0] or "").strip()
        name = str(values[1] or "").strip()
        category = str(values[2] or "").strip()
        if not code or not name or category != "Equity":
            continue
        rows.append({"symbol": f"{code}.HK", "code": code, "name": name, "market": "HK", "source": str(path)})
    return rows


def load_aliases(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle) if row.get("query") and row.get("symbol")]


def candidate_from_alias(query: str, aliases: list[dict[str, str]]) -> list[Candidate]:
    normalized_query = normalize_text(query)
    candidates: list[Candidate] = []
    for row in aliases:
        alias = normalize_text(row.get("query", ""))
        name = row.get("name", "").strip() or row.get("query", "").strip()
        if normalized_query == alias:
            confidence = float(row.get("confidence") or 0.98)
            match_type = "alias_exact"
        elif normalized_query and (normalized_query in alias or alias in normalized_query):
            confidence = float(row.get("partial_confidence") or 0.86)
            match_type = "alias_partial"
        else:
            continue
        candidates.append(
            Candidate(
                symbol=normalize_code(row.get("symbol", "")),
                name=name,
                market=(row.get("market") or "alias").strip(),
                confidence=confidence,
                match_type=match_type,
                source=str(DEFAULT_ALIAS_PATH),
            )
        )
    return candidates


def score_row(query: str, row: dict[str, str]) -> Candidate | None:
    raw_query = query.strip()
    normalized_query = normalize_text(raw_query)
    symbol = normalize_code(row["symbol"])
    code = normalize_code(row.get("code", ""))
    name = row["name"].strip()
    normalized_name = normalize_text(name)

    if not normalized_query:
        return None
    if normalize_code(raw_query) in {symbol, code}:
        return Candidate(symbol, name, row["market"], 1.0, "code_exact", row["source"])
    if normalized_query == normalized_name:
        return Candidate(symbol, name, row["market"], 0.98, "name_exact", row["source"])
    if normalized_query in normalized_name or normalized_name in normalized_query:
        return Candidate(symbol, name, row["market"], 0.88, "name_contains", row["source"])
    ratio = SequenceMatcher(None, normalized_query, normalized_name).ratio()
    if ratio >= 0.72:
        return Candidate(symbol, name, row["market"], round(0.55 + ratio * 0.3, 3), "name_fuzzy", row["source"])
    return None


def market_allowed(row_market: str, requested: str) -> bool:
    if requested == "all":
        return True
    if requested == "US":
        return row_market in {"NASDAQ", "NYSE", "AMEX", "US"}
    return row_market == requested


def dedupe(candidates: Iterable[Candidate]) -> list[Candidate]:
    best: dict[str, Candidate] = {}
    for candidate in candidates:
        current = best.get(candidate.symbol)
        if current is None or candidate.confidence > current.confidence:
            best[candidate.symbol] = candidate
    return sorted(best.values(), key=lambda item: (-item.confidence, item.market, item.symbol))


def query_symbols(query: str, *, market: str, limit: int, root: Path, alias_path: Path) -> dict[str, object]:
    rows = load_a_share(root) + load_hk(root) + load_us(root)
    candidates: list[Candidate] = []
    candidates.extend(candidate_from_alias(query, load_aliases(alias_path)))
    for row in rows:
        if not market_allowed(row["market"], market):
            continue
        candidate = score_row(query, row)
        if candidate:
            candidates.append(candidate)

    ranked = dedupe(candidate for candidate in candidates if market_allowed(candidate.market, market))[:limit]
    top = ranked[0] if ranked else None
    runner_up = ranked[1] if len(ranked) > 1 else None
    confirmed = bool(top and top.confidence >= 0.92 and (runner_up is None or top.confidence - runner_up.confidence >= 0.05))
    status = "confirmed" if confirmed else ("ambiguous" if ranked else "not_found")
    return {
        "query": query,
        "market": market,
        "status": status,
        "confirmed": confirmed,
        "recommendation": asdict(top) if top and confirmed else None,
        "candidates": [asdict(candidate) for candidate in ranked],
    }


def print_text(payload: dict[str, object]) -> None:
    print(f"查询: {payload['query']}")
    print(f"状态: {payload['status']}")
    recommendation = payload.get("recommendation")
    if recommendation:
        item = recommendation
        print(f"建议写入: {item['name']}({item['symbol']})")
    for item in payload["candidates"]:  # type: ignore[index]
        print(
            f"- {item['symbol']} | {item['name']} | {item['market']} | "
            f"{item['confidence']:.2f} | {item['match_type']}"
        )
    if payload["status"] != "confirmed":
        print("处理建议: 不要直接写入代码；放入“存疑与待确认”或补充人工校对参考。")


def main() -> int:
    parser = argparse.ArgumentParser(description="本地查询股票代码候选和置信度")
    parser.add_argument("query", help="公司名、简称、股票代码或 ticker")
    parser.add_argument("--market", choices=["all", "A", "HK", "US", "NASDAQ", "NYSE", "AMEX"], default="all")
    parser.add_argument("--limit", type=int, default=8)
    parser.add_argument("--root", default=str(DEFAULT_SYMBOL_ROOT), help="market-symbols 目录")
    parser.add_argument("--aliases", default=str(DEFAULT_ALIAS_PATH), help="公司别名 CSV")
    parser.add_argument("--json", action="store_true", help="输出 JSON")
    args = parser.parse_args()

    payload = query_symbols(
        args.query,
        market=args.market,
        limit=args.limit,
        root=Path(args.root).expanduser(),
        alias_path=Path(args.aliases).expanduser(),
    )
    if args.json:
        print(json.dumps(payload, ensure_ascii=False, indent=2))
    else:
        print_text(payload)
    return 0 if payload["candidates"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
